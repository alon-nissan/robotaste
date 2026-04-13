"""
Generate protocols/protocol_template.xlsx — a non-technical fill-in template
that can be sent to an LLM to produce a valid RoboTaste protocol JSON.

Run:
    python protocols/generate_template.py
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Palette
# ─────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark blue  → section headers
C_HEADER_FG   = "FFFFFF"   # white text
C_REQ_BG      = "FFF2CC"   # light yellow → required fields
C_OPT_BG      = "FFFFFF"   # white       → optional fields
C_EX_BG       = "F2F2F2"   # light grey  → example column
C_NOTE_BG     = "E2EFDA"   # light green → notes column
C_SUBHDR_BG   = "BDD7EE"   # medium blue → sub-section headers
C_TABLE_HDR   = "2E75B6"   # table column headers
C_ALT_ROW     = "DEEAF1"   # alternating table row


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, color="000000", italic=False, size=11) -> Font:
    return Font(bold=bold, color=color, italic=italic, size=size)


def border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def wrap_align(horizontal="left", wrap=True) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="top", wrap_text=wrap)


def yes_no_validation() -> DataValidation:
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    dv.error = "Enter yes or no"
    dv.errorTitle = "Invalid value"
    return dv


def write_row(ws, row: int, label: str, default="", example="", notes="",
              required=False, validation: DataValidation = None):
    """Write a single key-value row."""
    bg = C_REQ_BG if required else C_OPT_BG

    # A: label
    c = ws.cell(row=row, column=1, value=label)
    c.font = font(bold=required)
    c.fill = fill(bg)
    c.alignment = wrap_align()
    c.border = border_thin()

    # B: fill-in value
    c = ws.cell(row=row, column=2, value=default)
    c.fill = fill(bg)
    c.alignment = wrap_align()
    c.border = border_thin()
    if validation:
        validation.add(c)

    # C: example
    c = ws.cell(row=row, column=3, value=example)
    c.fill = fill(C_EX_BG)
    c.font = font(italic=True, color="595959")
    c.alignment = wrap_align()
    c.border = border_thin()

    # D: notes
    c = ws.cell(row=row, column=4, value=notes)
    c.fill = fill(C_NOTE_BG)
    c.font = font(size=10, color="375623")
    c.alignment = wrap_align()
    c.border = border_thin()


def write_section(ws, row: int, title: str, subtitle: str = "") -> int:
    """Write a coloured section header. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=title)
    c.font = font(bold=True, color=C_HEADER_FG, size=12)
    c.fill = fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border_thin()
    row += 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = font(italic=True, color="595959", size=10)
        c.fill = fill(C_SUBHDR_BG)
        c.alignment = wrap_align()
        c.border = border_thin()
        row += 1
    return row


def write_col_headers(ws, row: int, headers: list) -> int:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()
    return row + 1


def write_empty_table_rows(ws, start_row: int, num_rows: int,
                           num_cols: int, alt=False) -> int:
    for r in range(start_row, start_row + num_rows):
        bg = C_ALT_ROW if (alt and r % 2 == 0) else C_OPT_BG
        for col in range(1, num_cols + 1):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(bg)
            c.alignment = wrap_align()
            c.border = border_thin()
    return start_row + num_rows


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════════

def build_readme(wb: Workbook):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 90

    lines = [
        ("RoboTaste Protocol Template", True, 16, C_HEADER_BG, C_HEADER_FG),
        ("", False, 11, C_OPT_BG, "000000"),
        ("HOW TO USE THIS TEMPLATE", True, 13, C_SUBHDR_BG, "1F4E79"),
        ("1.  Fill in each sheet (Basic Info → Screens → Settings).", False, 11, C_OPT_BG, "000000"),
        ("2.  Yellow rows are REQUIRED. White/grey rows are optional.", False, 11, C_OPT_BG, "000000"),
        ("3.  Drop-down cells have validation — click and choose from the list.", False, 11, C_OPT_BG, "000000"),
        ("4.  In table sheets, fill one row per ingredient / pump / sample / question.", False, 11, C_OPT_BG, "000000"),
        ("5.  Leave rows blank if not needed (do not delete them).", False, 11, C_OPT_BG, "000000"),
        ("", False, 11, C_OPT_BG, "000000"),
        ("SHEETS IN THIS WORKBOOK", True, 13, C_SUBHDR_BG, "1F4E79"),
        ("  Basic Info     — protocol name, description, author, phase sequence", False, 11, C_OPT_BG, "000000"),
        ("  Ingredients    — list of taste compounds (up to 6)", False, 11, C_OPT_BG, "000000"),
        ("  Pumps          — hardware pump configuration (skip if no pumps)", False, 11, C_OPT_BG, "000000"),
        ("  Sample Schedule— how samples are chosen per cycle block", False, 11, C_OPT_BG, "000000"),
        ("  Sample Bank    — fixed set of samples for predetermined_randomized mode", False, 11, C_OPT_BG, "000000"),
        ("  Questionnaire  — rating questions shown to subjects after each taste", False, 11, C_OPT_BG, "000000"),
        ("  Screens        — instructions, consent, and loading screen text", False, 11, C_OPT_BG, "000000"),
        ("  Settings       — stopping criteria, BO algorithm, data collection", False, 11, C_OPT_BG, "000000"),
        ("", False, 11, C_OPT_BG, "000000"),
        ("SENDING TO AN LLM", True, 13, C_SUBHDR_BG, "1F4E79"),
        ("After filling in the template, attach this file (or paste the contents) to", False, 11, C_OPT_BG, "000000"),
        ("your LLM chat and use the prompt below:", False, 11, C_OPT_BG, "000000"),
        ("", False, 11, C_OPT_BG, "000000"),
        (
            '┌─ COPY THIS PROMPT ──────────────────────────────────────────────────────────────┐\n'
            '│                                                                                   │\n'
            '│  I have filled in a RoboTaste protocol template (attached).                       │\n'
            '│  Please convert it into a valid RoboTaste protocol JSON file.                     │\n'
            '│                                                                                   │\n'
            '│  Rules:                                                                           │\n'
            '│  • Use exactly the field names and structure from the RoboTaste JSON schema.      │\n'
            '│  • Generate a unique protocol_id string (e.g. "proto_<name>_001").                │\n'
            '│  • Set created_at to today\'s date in ISO-8601 format.                            │\n'
            '│  • Convert all yes/no values to true/false.                                       │\n'
            '│  • Concentrations are in mM; volumes in mL or µL as specified.                   │\n'
            '│  • If a field is blank/optional and has a sensible default, use the default.      │\n'
            '│  • Output only the JSON — no explanation, no markdown fences.                    │\n'
            '│                                                                                   │\n'
            '└───────────────────────────────────────────────────────────────────────────────────┘',
            False, 10, "FFF9C4", "000000"
        ),
        ("", False, 11, C_OPT_BG, "000000"),
        ("COLOUR LEGEND", True, 13, C_SUBHDR_BG, "1F4E79"),
        ("  Yellow background   = REQUIRED field", False, 11, C_REQ_BG, "000000"),
        ("  White background    = optional field", False, 11, C_OPT_BG, "000000"),
        ("  Grey background     = example / hint (do not edit)", False, 11, C_EX_BG, "000000"),
        ("  Green background    = notes / valid options (do not edit)", False, 11, C_NOTE_BG, "000000"),
    ]

    for i, item in enumerate(lines, 1):
        text, bold, size, bg, fg = item
        ws.row_dimensions[i].height = 18 if "\n" not in text else 110
        c = ws.cell(row=i, column=1, value=text)
        c.font = font(bold=bold, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border_thin()


def build_basic_info(wb: Workbook):
    ws = wb.create_sheet("Basic Info")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 48

    yn = yes_no_validation()
    ws.add_data_validation(yn)

    phase_dv = DataValidation(
        type="list",
        formula1='"default (no consent),with consent"',
        allow_blank=True
    )
    ws.add_data_validation(phase_dv)

    # Column header row
    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "BASIC INFORMATION")
    write_row(ws, r, "Protocol Name", "", "Sucrose Dose Response", "Short, descriptive name (max 200 chars)", required=True); r+=1
    write_row(ws, r, "Description", "", "Dose-response study mapping sweetness perception across 6 sucrose levels", "What does this experiment do?"); r+=1
    write_row(ws, r, "Version", "1.0", "1.0", "major.minor format"); r+=1
    write_row(ws, r, "Created By", "", "Alon Nissan", "Researcher's name"); r+=1
    write_row(ws, r, "Tags (comma-separated)", "", "dose-response, sucrose, pilot", "Free-form labels for searching"); r+=1

    r += 1
    r = write_section(ws, r, "PHASE SEQUENCE",
                      "Which screens appear before the experiment? (choose from dropdown)")
    write_row(ws, r, "Phase Sequence", "with consent",
              "with consent",
              '"default (no consent)" = registration → instructions → experiment\n"with consent" = consent → registration → instructions → experiment',
              validation=phase_dv); r+=1

    r += 1
    r = write_section(ws, r, "DATA COLLECTION (optional)")
    write_row(ws, r, "Collect Demographics", "yes", "yes", "Ask age, gender etc. at registration", validation=yn); r+=1
    write_row(ws, r, "Track Interaction Times", "no", "no", "Record how long subjects take to answer", validation=yn); r+=1
    write_row(ws, r, "Track Trajectory", "no", "no", "Log slider movement during rating", validation=yn); r+=1


def build_ingredients(wb: Workbook):
    ws = wb.create_sheet("Ingredients")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20

    # Instructions banner
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value="List each ingredient (1–6 rows). Water used as diluent should have Is Diluent? = yes. "
                      "Available ingredients: Sugar, Salt, Citric Acid, Caffeine, MSG, Water.")
    c.font = font(italic=True, size=10)
    c.fill = fill(C_SUBHDR_BG)
    c.alignment = wrap_align()
    c.border = border_thin()
    ws.row_dimensions[1].height = 32

    headers = ["Ingredient Name *", "Min Conc (mM) *", "Max Conc (mM) *",
               "Stock Concentration (mM)", "Is Diluent?", "Notes"]
    write_col_headers(ws, 2, headers)

    # Example row
    examples = [["Sugar", 0.0, 100.0, 200.0, "no", ""],
                ["Water", 0.0, 0.0, 0.0, "yes", "Fills remaining volume"]]
    for i, ex in enumerate(examples):
        r = 3 + i
        for col, val in enumerate(ex, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(C_EX_BG)
            c.font = font(italic=True, color="595959")
            c.alignment = wrap_align()
            c.border = border_thin()

    # ── merge-cells note above example rows ──
    ws.merge_cells("A3:F3")
    # undo the merge — we actually want individual cells
    ws.unmerge_cells("A3:F3")
    # Re-write example rows properly (already done above)

    # Blank fill-in rows (rows 5-10)
    yn = yes_no_validation()
    ws.add_data_validation(yn)
    for r in range(5, 11):
        for col in range(1, 7):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(C_OPT_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        yn.add(ws.cell(row=r, column=5))

    ws.merge_cells("A3:F3")
    ws.unmerge_cells("A3:F3")

    # Note row
    ws.merge_cells("A12:F12")
    c = ws.cell(row=12, column=1,
                value="★  Rows 5-10 above are for YOUR ingredients. Grey rows 3-4 are examples only.")
    c.font = font(italic=True, size=10, color="375623")
    c.fill = fill(C_NOTE_BG)
    c.alignment = wrap_align()
    c.border = border_thin()


def build_pumps(wb: Workbook):
    ws = wb.create_sheet("Pumps")

    # Config section
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45

    yn = yes_no_validation()
    ws.add_data_validation(yn)

    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "PUMP CONTROL — skip this entire sheet if not using pumps")
    write_row(ws, r, "Enable Pumps", "no", "yes", "yes = robot dispenses samples automatically", validation=yn); r+=1
    write_row(ws, r, "Serial Port", "", "/dev/cu.PL2303G-USBtoUART120",
              "Mac/Linux: /dev/cu.XXX or /dev/ttyUSBX\nWindows: COM3"); r+=1
    write_row(ws, r, "Baud Rate", "19200", "19200", "19200 for NE-4000 pumps"); r+=1
    write_row(ws, r, "Total Sample Volume (mL)", "", "10", "mL dispensed per cycle"); r+=1
    write_row(ws, r, "Dispensing Rate (µL/min)", "", "90000", "Speed of dispensing"); r+=1
    write_row(ws, r, "Simultaneous Dispensing", "yes", "yes", "All pumps dispense at the same time", validation=yn); r+=1
    write_row(ws, r, "Burst Mode", "yes", "yes", "Faster network command (requires addresses 0-9)", validation=yn); r+=1

    r += 1
    r = write_section(ws, r, "INDIVIDUAL PUMP CONFIGURATION",
                      "One row per pump. Address = hardware address knob value.")

    pump_headers = ["Address *", "Ingredient *", "Syringe Diameter (mm) *",
                    "Max Rate (µL/min)", "Stock Conc (mM)", "Dual Syringe?",
                    "Tube Volume (µL)", "Purge Volume (µL)", "Notes"]
    for ci, h in enumerate(pump_headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = [8, 16, 22, 18, 16, 16, 16, 16, 28][ci-1]
    r += 1

    # Example rows
    ex_pumps = [
        [0, "Water", 29.0, 90000, 0.0, "yes", 1500, 1500, "BD 50mL syringes"],
        [1, "Sugar", 29.0, 90000, 200.0, "yes", 1500, 1500, "200 mM sucrose stock"],
    ]
    for ex in ex_pumps:
        for ci, val in enumerate(ex, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill(C_EX_BG)
            c.font = font(italic=True, color="595959")
            c.alignment = wrap_align()
            c.border = border_thin()
        r += 1

    # Blank fill-in rows
    yn2 = yes_no_validation()
    ws.add_data_validation(yn2)
    for _ in range(6):
        for ci in range(1, len(pump_headers)+1):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = fill(C_OPT_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        yn2.add(ws.cell(row=r, column=6))
        r += 1

    ws.merge_cells(f"A{r}:{get_column_letter(len(pump_headers))}{r}")
    note = ws.cell(row=r, column=1,
                   value="★  Grey rows above are examples. Fill in the blank rows below them.")
    note.font = font(italic=True, size=10, color="375623")
    note.fill = fill(C_NOTE_BG)
    note.alignment = wrap_align()
    note.border = border_thin()


def build_sample_schedule(wb: Workbook):
    ws = wb.create_sheet("Sample Schedule")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55

    mode_dv = DataValidation(
        type="list",
        formula1='"predetermined_randomized,predetermined_absolute,user_selected,bo_selected"',
        allow_blank=True
    )
    ws.add_data_validation(mode_dv)

    # Instructions
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1,
                value="Define cycle blocks. Each block has a cycle range and a selection mode. "
                      "Add more blocks if the experiment has different phases (e.g. warm-up then BO).")
    c.font = font(italic=True, size=10)
    c.fill = fill(C_SUBHDR_BG)
    c.alignment = wrap_align()
    c.border = border_thin()
    ws.row_dimensions[1].height = 36

    headers = ["Start Cycle *", "End Cycle *", "Mode *", "Notes"]
    write_col_headers(ws, 2, headers)

    # Mode description notes
    mode_notes = (
        "predetermined_randomized — samples drawn from a fixed Sample Bank (see that sheet)\n"
        "predetermined_absolute   — exact concentrations specified per cycle\n"
        "user_selected            — subject chooses their own concentrations\n"
        "bo_selected              — Bayesian Optimisation picks the next sample"
    )

    # Example row
    for ci, val in enumerate([1, 6, "predetermined_randomized", mode_notes], 1):
        c = ws.cell(row=3, column=ci, value=val)
        c.fill = fill(C_EX_BG)
        c.font = font(italic=True, color="595959")
        c.alignment = wrap_align()
        c.border = border_thin()
    ws.row_dimensions[3].height = 72

    # Blank fill-in rows (blocks 1-5)
    for r in range(4, 9):
        for ci in range(1, 5):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = fill(C_OPT_BG if r > 4 else C_REQ_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        mode_dv.add(ws.cell(row=r, column=3))

    ws.merge_cells("A10:D10")
    note = ws.cell(row=10, column=1,
                   value="★  Row 3 is an example. Row 4 (yellow) is your first block — required. "
                         "Rows 5-8 are optional additional blocks.")
    note.font = font(italic=True, size=10, color="375623")
    note.fill = fill(C_NOTE_BG)
    note.alignment = wrap_align()
    note.border = border_thin()

    # ── predetermined_absolute sub-section ──
    r = 12
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1,
                value="PREDETERMINED ABSOLUTE SAMPLES  (fill only if any block uses predetermined_absolute mode)")
    c.font = font(bold=True, color=C_HEADER_FG)
    c.fill = fill(C_HEADER_BG)
    c.alignment = wrap_align()
    c.border = border_thin()
    r += 1

    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1,
                value="One row per cycle. Ingredient columns must match the names on the Ingredients sheet.")
    c.font = font(italic=True, size=10)
    c.fill = fill(C_SUBHDR_BG)
    c.alignment = wrap_align()
    c.border = border_thin()
    r += 1

    for ci, h in enumerate(["Cycle #", "Ingredient 1 Conc (mM)", "Ingredient 2 Conc (mM)",
                             "Ingredient 3 Conc (mM)"], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()
    r += 1

    for _ in range(8):
        for ci in range(1, 5):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = fill(C_OPT_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        r += 1


def build_sample_bank(wb: Workbook):
    ws = wb.create_sheet("Sample Bank")

    # Config
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 44

    yn = yes_no_validation()
    ws.add_data_validation(yn)
    design_dv = DataValidation(
        type="list", formula1='"latin_square,randomized"', allow_blank=True
    )
    ws.add_data_validation(design_dv)

    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "SAMPLE BANK CONFIG",
                      "Only needed when any block uses 'predetermined_randomized' mode")
    write_row(ws, r, "Design Type", "latin_square", "latin_square",
              "latin_square — each participant sees each sample exactly once before repeats\n"
              "randomized   — fully random order",
              validation=design_dv); r+=1
    write_row(ws, r, "Prevent Consecutive Repeats", "yes", "yes",
              "Do not serve the same sample twice in a row", validation=yn); r+=1
    write_row(ws, r, "Ensure All Used Before Repeat", "yes", "yes",
              "Each sample must appear once per 'round' before any repeats", validation=yn); r+=1

    r += 1
    r = write_section(ws, r, "SAMPLE BANK — one row per sample",
                      "ID = single letter (A, B, C…). "
                      "Ingredient columns must match names on Ingredients sheet.")

    bank_headers = ["Sample ID *", "Label", "Ingredient 1 Conc (mM) *",
                    "Ingredient 2 Conc (mM)", "Ingredient 3 Conc (mM)", "Notes"]
    for ci, h in enumerate(bank_headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = [12, 22, 24, 24, 24, 28][ci-1]
    r += 1

    # Example rows
    ex_samples = [
        ["A", "No Sugar",        0.0,  "", "", ""],
        ["B", "Very Low Sugar",  20.0, "", "", ""],
        ["C", "Low Sugar",       40.0, "", "", ""],
        ["D", "Medium Sugar",    60.0, "", "", ""],
        ["E", "High Sugar",      80.0, "", "", ""],
        ["F", "Very High Sugar", 100.0,"", "", ""],
    ]
    for ex in ex_samples:
        for ci, val in enumerate(ex, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill(C_EX_BG)
            c.font = font(italic=True, color="595959")
            c.alignment = wrap_align()
            c.border = border_thin()
        r += 1

    # Blank fill-in rows
    for _ in range(12):
        for ci in range(1, len(bank_headers)+1):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = fill(C_OPT_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        r += 1

    ws.merge_cells(f"A{r}:{get_column_letter(len(bank_headers))}{r}")
    note = ws.cell(row=r, column=1,
                   value="★  Grey rows are examples. Fill in blank rows for YOUR samples.")
    note.font = font(italic=True, size=10, color="375623")
    note.fill = fill(C_NOTE_BG)
    note.alignment = wrap_align()
    note.border = border_thin()


def build_questionnaire(wb: Workbook):
    ws = wb.create_sheet("Questionnaire")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 50

    yn = yes_no_validation()
    ws.add_data_validation(yn)
    type_dv = DataValidation(
        type="list", formula1='"slider,dropdown"', allow_blank=True
    )
    ws.add_data_validation(type_dv)
    disp_dv = DataValidation(
        type="list", formula1='"slider_continuous,pillboxes"', allow_blank=True
    )
    ws.add_data_validation(disp_dv)
    transform_dv = DataValidation(
        type="list", formula1='"identity,log,normalize"', allow_blank=True
    )
    ws.add_data_validation(transform_dv)

    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "QUESTIONNAIRE META")
    write_row(ws, r, "Questionnaire Name", "", "Sweetness Intensity Scale"); r+=1
    write_row(ws, r, "Description", "", "9-point continuous sweetness rating"); r+=1

    r += 1
    r = write_section(ws, r, "QUESTIONS TABLE — one row per question")

    q_headers = ["Question ID *", "Type *", "Label (shown to subject) *",
                 "Help Text", "Min *", "Max *", "Default", "Step",
                 "Required?", "Display Type", "Scale Labels (format: value:label, …)"]
    for ci, h in enumerate(q_headers, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = \
            [20, 12, 34, 28, 8, 8, 10, 8, 12, 18, 44][ci-1]
    r += 1

    # Example row
    ex_q = ["sweetness_intensity", "slider",
            "How sweet was the sample you just tasted?",
            "Rate from 1 (not sweet) to 9 (intense)",
            1.0, 9.0, 5.0, 0.01, "yes", "slider_continuous",
            "1:Not Sweet at All, 3:Light, 5:Medium, 7:High, 9:Intense Sweetness"]
    for ci, val in enumerate(ex_q, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = fill(C_EX_BG)
        c.font = font(italic=True, color="595959")
        c.alignment = wrap_align()
        c.border = border_thin()
    r += 1

    # Blank fill-in rows (5 questions max)
    yn2 = yes_no_validation()
    ws.add_data_validation(yn2)
    type_dv2 = DataValidation(type="list", formula1='"slider,dropdown"', allow_blank=True)
    ws.add_data_validation(type_dv2)
    disp_dv2 = DataValidation(type="list", formula1='"slider_continuous,pillboxes"', allow_blank=True)
    ws.add_data_validation(disp_dv2)
    for _ in range(5):
        for ci in range(1, len(q_headers)+1):
            c = ws.cell(row=r, column=ci, value="")
            c.fill = fill(C_OPT_BG)
            c.alignment = wrap_align()
            c.border = border_thin()
        type_dv2.add(ws.cell(row=r, column=2))
        yn2.add(ws.cell(row=r, column=9))
        disp_dv2.add(ws.cell(row=r, column=10))
        r += 1

    r += 1
    r = write_section(ws, r, "BAYESIAN OPTIMISATION TARGET",
                      "Which question score should BO try to optimise? (required if using bo_selected mode)")
    write_row(ws, r, "Target Question ID", "", "sweetness_intensity",
              "Must match a Question ID above"); r+=1
    write_row(ws, r, "Higher Score is Better?", "yes", "yes",
              "yes = maximise; no = minimise", validation=yn); r+=1
    write_row(ws, r, "Transform", "identity", "identity",
              "identity / log / normalize", validation=transform_dv); r+=1
    write_row(ws, r, "Expected Range (min, max)", "", "1, 9",
              "Expected min and max of the question score"); r+=1
    write_row(ws, r, "Optimal Threshold", "", "7.0",
              "Score considered 'good enough' (used in some stopping criteria"); r+=1


def build_screens(wb: Workbook):
    ws = wb.create_sheet("Screens")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 46

    yn = yes_no_validation()
    ws.add_data_validation(yn)
    size_dv = DataValidation(
        type="list", formula1='"normal,large,extra_large"', allow_blank=True
    )
    ws.add_data_validation(size_dv)

    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "CONSENT FORM (shown before experiment if phase sequence includes consent)")
    consent_text = (
        "Dear Participant,\n\n"
        "Welcome to the RoboTaste experiment.\n"
        "Participation will take approximately 10 minutes.\n"
        "There are no right or wrong answers.\n"
        "Your answers are confidential.\n"
        "Participation is voluntary and you may stop at any time."
    )
    write_row(ws, r, "Consent Explanation Text", "", consent_text,
              "Shown to participant. Supports plain text. Use \\n for line breaks."); r+=1
    write_row(ws, r, "Contact Information", "",
              "For questions contact: Prof. X (x@university.edu)",
              "Shown at the bottom of the consent page"); r+=1
    write_row(ws, r, "Consent Checkbox Label", "",
              "I have read the above information and agree to participate.",
              "Text next to the tick-box the participant must check"); r+=1

    r += 1
    r = write_section(ws, r, "INSTRUCTIONS SCREEN (shown after consent / before experiment)")
    inst_text = (
        "**Welcome!**\n\n"
        "In this experiment you will taste liquid samples.\n\n"
        "For each sample:\n"
        "1. The robot will prepare your sample.\n"
        "2. Take the cup and place a clean cup under the spout.\n"
        "3. Rinse your mouth and stir the sample lightly.\n"
        "4. Rate the taste on the scale provided.\n"
        "5. Rinse your mouth before the next sample."
    )
    write_row(ws, r, "Instructions Title", "Instructions", "Instructions",
              "Heading shown on the instructions page"); r+=1
    write_row(ws, r, "Instructions Text (markdown)", "", inst_text,
              "Supports **bold**, *italic*, numbered lists. Use \\n for new lines."); r+=1
    write_row(ws, r, "Confirm Checkbox Label", "",
              "I have read and understand the instructions.",
              "Text next to the checkbox at the bottom"); r+=1
    write_row(ws, r, "Begin Button Label", "Start Tasting", "Begin Experiment",
              "Label on the proceed button"); r+=1

    r += 1
    r = write_section(ws, r, "LOADING SCREEN (shown between each sample while robot prepares)")
    loading_text = (
        "Please rinse your mouth with water before and after every tasting.\n\n"
        "Once the robot is done, take the cup, stir lightly, "
        "and place a clean cup under the spout."
    )
    write_row(ws, r, "Loading Message", "", loading_text,
              "Reminder instructions shown during preparation"); r+=1
    write_row(ws, r, "Duration (seconds)", "12", "12",
              "How long to show the screen (ignored if dynamic duration is used)"); r+=1
    write_row(ws, r, "Use Dynamic Duration", "yes", "yes",
              "yes = wait for pump to finish (recommended with pumps)", validation=yn); r+=1
    write_row(ws, r, "Show Progress Bar", "yes", "yes", "", validation=yn); r+=1
    write_row(ws, r, "Show Cycle Info", "yes", "yes",
              "Display 'Sample 3 of 6' during loading", validation=yn); r+=1
    write_row(ws, r, "Message Size", "large", "large",
              "normal / large / extra_large", validation=size_dv); r+=1

    # Taller rows for text areas
    for tall_r in [4, 7, 13]:
        ws.row_dimensions[tall_r].height = 90


def build_settings(wb: Workbook):
    ws = wb.create_sheet("Settings")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 50

    yn = yes_no_validation()
    ws.add_data_validation(yn)
    acq_dv = DataValidation(
        type="list", formula1='"ei,ucb"', allow_blank=True
    )
    ws.add_data_validation(acq_dv)
    stop_dv = DataValidation(
        type="list",
        formula1='"manual_only,suggest_auto,auto_with_minimum"',
        allow_blank=True
    )
    ws.add_data_validation(stop_dv)

    for col, h in enumerate(["Field", "YOUR VALUE", "Example", "Notes"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font(bold=True, color=C_HEADER_FG)
        c.fill = fill(C_TABLE_HDR)
        c.alignment = wrap_align("center")
        c.border = border_thin()

    r = 2
    r = write_section(ws, r, "STOPPING CRITERIA")
    write_row(ws, r, "Minimum Cycles", "", "6",
              "Experiment will not end before this many cycles", required=True); r+=1
    write_row(ws, r, "Maximum Cycles", "", "6",
              "Hard limit on cycles", required=True); r+=1
    write_row(ws, r, "Stop Mode", "manual_only", "manual_only",
              "manual_only — moderator stops manually\n"
              "suggest_auto — BO suggests when to stop\n"
              "auto_with_minimum — BO stops automatically after min_cycles",
              validation=stop_dv); r+=1

    r += 1
    r = write_section(ws, r, "BAYESIAN OPTIMISATION (skip if not using bo_selected mode)")
    write_row(ws, r, "Enable BO", "no", "yes",
              "Set to yes only if a block uses bo_selected mode", validation=yn); r+=1
    write_row(ws, r, "Acquisition Function", "ei", "ei",
              "ei = Expected Improvement (default)\n"
              "ucb = Upper Confidence Bound",
              validation=acq_dv); r+=1
    write_row(ws, r, "Min Samples Before BO Starts", "3", "3",
              "BO waits for this many observations before making suggestions"); r+=1
    write_row(ws, r, "EI Xi (exploration trade-off)", "0.01", "0.01",
              "For EI: 0.01 = exploit, 0.1 = explore. Ignored for UCB."); r+=1
    write_row(ws, r, "UCB Kappa", "2.5", "2.5",
              "For UCB: higher = more exploration. Ignored for EI."); r+=1
    write_row(ws, r, "Adaptive Acquisition", "yes", "yes",
              "Automatically switch between explore/exploit over time", validation=yn); r+=1


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    build_readme(wb)
    build_basic_info(wb)
    build_ingredients(wb)
    build_pumps(wb)
    build_sample_schedule(wb)
    build_sample_bank(wb)
    build_questionnaire(wb)
    build_screens(wb)
    build_settings(wb)

    out = Path(__file__).parent / "protocol_template.xlsx"
    wb.save(out)
    print(f"✓  Written: {out}")


if __name__ == "__main__":
    main()
